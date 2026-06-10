"""WP-80 — v2 corpus handling: skip categories, no-extension sniffing, archive
expansion, content-hash dedup. Fully offline and deterministic."""
from __future__ import annotations

import os
import zipfile

os.environ.setdefault("MTA_AUTO_UPDATE", "off")

from mta.core.config import Config
from mta.core.convert import convert_file


def _cfg(tmp_path, **over):
    c = Config(home=tmp_path / "home")
    for k, v in over.items():
        setattr(c, k, v)
    return c


# ---- skip categories (never read; status='skipped') ---------------------------------
def test_skip_categories(tmp_path):
    cases = {
        "photo.jpg": "media", "clip.mp4": "media", "talk.mp3": "media",
        "Sutonny.ttf": "font", "brand.otf": "font",
        "policy.pdf.gdrive": "gdrive-pointer", "sheet.gsheet": "gdrive-pointer",
        "scratch.tmp": "junk", ".DS_Store": "junk", "Thumbs.db": "junk",
    }
    for name, cat in cases.items():
        f = tmp_path / name
        f.write_bytes(b"x" * 32)            # content is irrelevant — name-based skip
        res = convert_file(f, tmp_path / "out", _cfg(tmp_path))
        assert res.status == "skipped" and res.method == "skipped-type", (name, res.status)
        assert res.error == cat, (name, res.error, cat)


def test_skip_switches_can_be_disabled(tmp_path):
    f = tmp_path / "photo.png"
    f.write_bytes(b"\x89PNG\r\n\x1a\n" + b"\x00" * 64)   # real-ish PNG header, no OCR dep
    res = convert_file(f, tmp_path / "out", _cfg(tmp_path, skip_media=False, ocr_mode="off"))
    # With media skipping off and OCR off, an image converts to nothing → empty/unsupported,
    # but it must NOT be 'skipped-type' (the switch was honoured).
    assert res.method != "skipped-type"


def test_text_files_still_convert(tmp_path):
    f = tmp_path / "note.txt"
    f.write_text("Aurora budget approved.", encoding="utf-8")
    res = convert_file(f, tmp_path / "out", _cfg(tmp_path))
    assert res.status == "ok"


# ---- no-extension Office/zip sniffing ------------------------------------------------
def test_no_extension_xlsx_is_sniffed(tmp_path):
    # Build a minimal real .xlsx via openpyxl if available, else fall back to checking
    # that a PK-magic file at least routes through the sniff path (not 'binary').
    target = tmp_path / "Rural data (Missining 1)"   # the real corpus case: no extension
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Helios cohort"
        ws["B1"] = 42
        wb.save(target)
    except ImportError:
        # minimal zip with PK magic — MarkItDown will fail it, but it must not be
        # classified as plain binary without trying.
        with zipfile.ZipFile(target, "w") as z:
            z.writestr("content.txt", "hello")
    res = convert_file(target, tmp_path / "out", _cfg(tmp_path))
    import importlib.util
    if importlib.util.find_spec("markitdown") is not None:
        # With a converter present, the PK-magic file is sniffed and read as OOXML.
        assert res.status == "ok" and "sniffed-zip" in res.method, (res.status, res.method)
    else:
        # No converter installed (CI's minimal offline env): a zip can't be read, so a
        # clean terminal status is correct — the point is it never CRASHES and isn't
        # mis-handled as a missing file.
        assert res.status in ("unsupported", "empty"), (res.status, res.method)


# ---- recursive archive expansion (security-critical) --------------------------------
def _expand_with(tmp_path, cfg, paths):
    from mta.core.digest import _expand
    return _expand([str(p) for p in paths], cfg=cfg)


def test_zip_expands_and_members_digest(tmp_path):
    src = tmp_path / "in"; src.mkdir()
    z = src / "reports.zip"
    with zipfile.ZipFile(z, "w") as zf:
        zf.writestr("a/report1.txt", "Helios cohort graduated.")
        zf.writestr("report2.txt", "Borealis funding approved.")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    names = {f.name for f in files}
    assert "report1.txt" in names and "report2.txt" in names
    assert "reports.zip" not in names                      # replaced by its members


def test_nested_zip_expands_recursively(tmp_path):
    import io
    src = tmp_path / "in"; src.mkdir()
    inner = io.BytesIO()
    with zipfile.ZipFile(inner, "w") as zi:
        zi.writestr("deep.txt", "Nested fact: Sherpur dataset complete.")
    outer = src / "outer.zip"
    with zipfile.ZipFile(outer, "w") as zo:
        zo.writestr("inner.zip", inner.getvalue())
        zo.writestr("top.txt", "Top-level note.")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    names = {f.name for f in files}
    assert "deep.txt" in names and "top.txt" in names
    assert "inner.zip" not in names                        # nested archive consumed


def test_zip_slip_member_is_rejected(tmp_path):
    src = tmp_path / "in"; src.mkdir()
    z = src / "evil.zip"
    with zipfile.ZipFile(z, "w") as zf:
        zf.writestr("../escape.txt", "I escaped!")
        zf.writestr("good.txt", "legit")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    assert any(f.name == "good.txt" for f in files)        # good member survives
    assert not (tmp_path / "escape.txt").exists()          # traversal blocked
    assert not (cfg.unpack_dir.parent / "escape.txt").exists()


def test_tar_symlink_member_is_skipped(tmp_path):
    import io
    import tarfile as tf
    src = tmp_path / "in"; src.mkdir()
    t = src / "tricky.tar"
    with tf.open(t, "w") as tar:
        info = tf.TarInfo("link.txt"); info.type = tf.SYMTYPE; info.linkname = "/etc/passwd"
        tar.addfile(info)
        data = b"safe content"
        info2 = tf.TarInfo("safe.txt"); info2.size = len(data)
        tar.addfile(info2, io.BytesIO(data))
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    names = {f.name for f in files}
    assert "safe.txt" in names and "link.txt" not in names


def test_archive_depth_cap_terminates(tmp_path):
    import io
    src = tmp_path / "in"; src.mkdir()
    payload = io.BytesIO()
    with zipfile.ZipFile(payload, "w") as z:
        z.writestr("core.txt", "innermost")
    blob = payload.getvalue()
    for i in range(6):                                     # 6 layers of nesting
        outer = io.BytesIO()
        with zipfile.ZipFile(outer, "w") as z:
            z.writestr(f"layer{i}.zip", blob)
        blob = outer.getvalue()
    deep = src / "deep.zip"; deep.write_bytes(blob)
    cfg = _cfg(tmp_path, archive_max_depth=3)
    files = _expand_with(tmp_path, cfg, [src])             # must terminate, never loop
    assert isinstance(files, list)                         # no crash; depth-capped


def test_archive_entry_cap_rejects_whole_archive(tmp_path):
    src = tmp_path / "in"; src.mkdir()
    z = src / "many.zip"
    with zipfile.ZipFile(z, "w") as zf:
        for i in range(10):
            zf.writestr(f"f{i}.txt", f"file {i}")
    cfg = _cfg(tmp_path, archive_max_entries=3)
    files = _expand_with(tmp_path, cfg, [src])
    assert any(f.name == "many.zip" for f in files)        # kept → honest 'skipped'
    assert not any(f.name == "f1.txt" for f in files)      # nothing half-ingested
    res = convert_file(z, tmp_path / "out", cfg)
    assert res.status == "skipped" and res.method == "archive"


def test_targz_and_single_stream_gz(tmp_path):
    import tarfile as tf
    src = tmp_path / "in"; src.mkdir()
    inner = tmp_path / "doc.txt"; inner.write_text("tar.gz payload", encoding="utf-8")
    with tf.open(src / "bundle.tar.gz", "w:gz") as tar:
        tar.add(inner, arcname="doc.txt")
    import gzip as gz
    with gz.open(src / "notes.txt.gz", "wb") as f:
        f.write(b"gzip payload")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    names = {f.name for f in files}
    assert "doc.txt" in names and "notes.txt" in names


def test_corrupt_rar_is_handled_cleanly(tmp_path):
    # Invariant (regardless of whether unar/7z is installed): a corrupt rar never
    # crashes _expand, never injects garbage members, and — if it's kept (no extractor,
    # or the extractor declined it) — convert_file reports a clean 'skipped'.
    src = tmp_path / "in"; src.mkdir()
    bad = src / "broken.rar"; bad.write_bytes(b"Rar!\x1a\x07\x00garbagegarbage")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])              # must not raise
    assert not any("garbage" in f.name for f in files)     # no junk member leaked
    if any(f.name == "broken.rar" for f in files):
        res = convert_file(bad, tmp_path / "out", cfg)
        assert res.status == "skipped" and res.method == "archive"


def test_content_hash_dedup(tmp_path):
    src = tmp_path / "in"; src.mkdir()
    (src / "a.txt").write_text("identical content", encoding="utf-8")
    (src / "b_copy.txt").write_text("identical content", encoding="utf-8")
    (src / "c.txt").write_text("different content", encoding="utf-8")
    cfg = _cfg(tmp_path)
    files = _expand_with(tmp_path, cfg, [src])
    assert len(files) == 2                                  # one of the twins dropped
    assert any(f.name == "c.txt" for f in files)


# ---- legacy binary Office (LibreOffice fallback, v2.0.1) -----------------------------
def test_legacy_office_helper_degrades_without_libreoffice(tmp_path, monkeypatch):
    # Without LibreOffice on PATH, a .doc must degrade cleanly (never crash).
    from mta.core import convert as cv
    monkeypatch.setattr(cv, "_soffice_bin", lambda: None)
    f = tmp_path / "old.doc"
    f.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)   # OLE2 magic
    res = cv.convert_file(f, tmp_path / "out", _cfg(tmp_path))
    # .doc is in the markitdown set → routed to legacy handler → libreoffice-missing →
    # markitdown can't read it either → a clean terminal status, not a crash.
    assert res.status in ("unsupported", "failed", "empty"), (res.status, res.method)


def test_legacy_office_converts_when_soffice_present(tmp_path):
    import shutil
    import pytest
    from mta.core.convert import _soffice_bin, convert_file
    if _soffice_bin() is None:
        pytest.skip("LibreOffice not installed")
    # Build a real .doc by converting a .docx (needs python-docx); else skip.
    try:
        import docx
    except ImportError:
        pytest.skip("python-docx not installed to build a .doc fixture")
    d = docx.Document()
    d.add_paragraph("Dr. Karim Rahman leads Project Aurora for the Nordic Grid Authority.")
    src_docx = tmp_path / "src.docx"; d.save(src_docx)
    # LibreOffice-convert docx → doc to get a genuine legacy binary fixture
    import subprocess
    subprocess.run([_soffice_bin(), "--headless", "--convert-to", "doc",
                    "--outdir", str(tmp_path), str(src_docx)], capture_output=True, timeout=240)
    doc = tmp_path / "src.doc"
    if not doc.exists():
        pytest.skip("could not build a .doc fixture")
    res = convert_file(doc, tmp_path / "out", _cfg(tmp_path))
    assert res.status == "ok" and "soffice+" in res.method, (res.status, res.method)
    assert "Karim Rahman" in (tmp_path / "out").glob("*.md").__next__().read_text(encoding="utf-8")
